"""Analytics views — dashboard metrics, funnel, scores, timeline."""
from drf_spectacular.utils import extend_schema, OpenApiParameter, inline_serializer
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers as drf_serializers
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework.request import Request

from common.authentication import JWTRequestAuthentication
from common.permissions import require_permission

from .services import get_dashboard_metrics, get_funnel_data, get_score_distribution, get_timeline_data


@extend_schema(
    tags=["Analytics"],
    summary="Dashboard metrics",
    description="Returns high-level recruitment metrics for the current tenant.",
    responses={200: inline_serializer("DashboardMetrics", fields={
        "total_jobs_open": drf_serializers.IntegerField(),
        "total_applications": drf_serializers.IntegerField(),
        "total_calls_completed": drf_serializers.IntegerField(),
        "avg_candidate_score": drf_serializers.FloatField(allow_null=True),
        "applications_today": drf_serializers.IntegerField(),
        "calls_today": drf_serializers.IntegerField(),
        "shortlisted_count": drf_serializers.IntegerField(),
        "offers_count": drf_serializers.IntegerField(),
        "hiring_conversion_rate": drf_serializers.FloatField(),
    })},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.analytics.view")])
def dashboard(request: Request) -> Response:
    data = get_dashboard_metrics(str(request.tenant_id))
    return Response(data)


@extend_schema(
    tags=["Analytics"],
    summary="Application funnel",
    description="Returns application counts grouped by status — shows the recruitment funnel.",
    responses={200: inline_serializer("FunnelItem", fields={
        "status": drf_serializers.CharField(),
        "count": drf_serializers.IntegerField(),
    }, many=True)},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.analytics.view")])
def funnel(request: Request) -> Response:
    data = get_funnel_data(str(request.tenant_id))
    return Response(data)


@extend_schema(
    tags=["Analytics"],
    summary="Score distribution",
    description="Returns candidate score distribution in 10-point buckets (0-10, 10-20, ..., 90-100).",
    responses={200: inline_serializer("ScoreBucket", fields={
        "range": drf_serializers.CharField(),
        "count": drf_serializers.IntegerField(),
    }, many=True)},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.analytics.view")])
def scores(request: Request) -> Response:
    data = get_score_distribution(str(request.tenant_id))
    return Response(data)


@extend_schema(
    tags=["Analytics"],
    summary="Activity timeline",
    description="Returns daily counts of applications, completed calls, and hires over a time period.",
    parameters=[
        OpenApiParameter("period", OpenApiTypes.STR, enum=["7d", "30d", "90d"],
            description="Time period to include (default: 30d)"),
    ],
    responses={200: inline_serializer("TimelinePoint", fields={
        "date": drf_serializers.DateField(),
        "applications": drf_serializers.IntegerField(),
        "calls": drf_serializers.IntegerField(),
        "hires": drf_serializers.IntegerField(),
    }, many=True)},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.analytics.view")])
def timeline(request: Request) -> Response:
    period = request.query_params.get("period", "30d")
    data = get_timeline_data(str(request.tenant_id), period)
    return Response(data)


@extend_schema(
    tags=["Analytics"],
    summary="Export analytics report",
    description=(
        "Export a recruitment analytics report as CSV or Excel. "
        "Includes: funnel breakdown, score distribution, and timeline data. "
        "Use `report` param to select a specific report or get all combined."
    ),
    parameters=[
        OpenApiParameter("export_format", OpenApiTypes.STR, enum=["csv", "xlsx"], description="Export format (default: csv)"),
        OpenApiParameter("report", OpenApiTypes.STR, enum=["funnel", "scores", "timeline", "all"],
            description="Which report to export (default: all)"),
        OpenApiParameter("period", OpenApiTypes.STR, enum=["7d", "30d", "90d"],
            description="Time period for timeline report (default: 30d)"),
    ],
    responses={200: None},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.analytics.view")])
def export_report(request: Request):
    """Export analytics data as CSV or Excel."""
    from common.export import build_csv_response, build_excel_response

    tenant_id = str(request.tenant_id)
    export_format = request.query_params.get("export_format", "csv").lower()
    report_type = request.query_params.get("report", "all").lower()
    period = request.query_params.get("period", "30d")

    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    if report_type == "funnel":
        rows = get_funnel_data(tenant_id)
        columns = [("status", "Status"), ("count", "Count")]
        filename = f"funnel_report_{timestamp}"
    elif report_type == "scores":
        rows = get_score_distribution(tenant_id)
        columns = [("range", "Score Range"), ("count", "Count")]
        filename = f"score_distribution_{timestamp}"
    elif report_type == "timeline":
        rows = get_timeline_data(tenant_id, period)
        columns = [
            ("date", "Date"),
            ("applications", "Applications"),
            ("calls", "Calls Completed"),
            ("hires", "Hires"),
        ]
        filename = f"timeline_report_{period}_{timestamp}"
    else:
        # Combined "all" report — export funnel as the primary sheet/data
        # with dashboard metrics as a header section
        metrics = get_dashboard_metrics(tenant_id)
        funnel = get_funnel_data(tenant_id)
        timeline_data = get_timeline_data(tenant_id, period)

        if export_format == "xlsx":
            return _build_combined_excel_report(
                metrics=metrics,
                funnel=funnel,
                timeline_data=timeline_data,
                timestamp=timestamp,
                period=period,
            )

        # CSV: export funnel + timeline combined
        rows = []
        rows.append({"section": "--- DASHBOARD METRICS ---", "metric": "", "value": ""})
        for key, val in metrics.items():
            rows.append({"section": "", "metric": key, "value": str(val) if val is not None else ""})
        rows.append({"section": "", "metric": "", "value": ""})
        rows.append({"section": "--- FUNNEL ---", "metric": "", "value": ""})
        for item in funnel:
            rows.append({"section": "", "metric": item["status"], "value": str(item["count"])})
        rows.append({"section": "", "metric": "", "value": ""})
        rows.append({"section": f"--- TIMELINE ({period}) ---", "metric": "", "value": ""})
        for item in timeline_data:
            rows.append({
                "section": item["date"],
                "metric": f"Apps:{item['applications']} Calls:{item['calls']} Hires:{item['hires']}",
                "value": "",
            })

        columns = [("section", "Section/Date"), ("metric", "Metric"), ("value", "Value")]
        filename = f"analytics_report_{timestamp}"

    if export_format == "xlsx":
        return build_excel_response(
            rows=rows,
            columns=columns,
            filename=f"{filename}.xlsx",
            sheet_name=report_type.title(),
        )
    return build_csv_response(
        rows=rows,
        columns=columns,
        filename=f"{filename}.csv",
    )


def _build_combined_excel_report(*, metrics, funnel, timeline_data, timestamp, period):
    """Build a multi-sheet Excel workbook with all analytics data."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    label_font = Font(bold=True)

    # Sheet 1: Dashboard Metrics
    ws = wb.active
    ws.title = "Dashboard"
    ws.cell(row=1, column=1, value="Metric").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value="Value").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    for idx, (key, val) in enumerate(metrics.items(), start=2):
        ws.cell(row=idx, column=1, value=key).font = label_font
        ws.cell(row=idx, column=2, value=str(val) if val is not None else "")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Sheet 2: Funnel
    ws2 = wb.create_sheet("Funnel")
    ws2.cell(row=1, column=1, value="Status").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Count").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    for idx, item in enumerate(funnel, start=2):
        ws2.cell(row=idx, column=1, value=item["status"])
        ws2.cell(row=idx, column=2, value=item["count"])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    # Sheet 3: Timeline
    ws3 = wb.create_sheet(f"Timeline ({period})")
    headers = ["Date", "Applications", "Calls Completed", "Hires"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for idx, item in enumerate(timeline_data, start=2):
        ws3.cell(row=idx, column=1, value=item["date"])
        ws3.cell(row=idx, column=2, value=item["applications"])
        ws3.cell(row=idx, column=3, value=item["calls"])
        ws3.cell(row=idx, column=4, value=item["hires"])
    for col_idx in range(1, 5):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="analytics_report_{timestamp}.xlsx"'
    return response
