"""Reusable CSV and Excel export utilities."""
import csv
import io
from datetime import datetime
from typing import Any, Sequence

from django.http import HttpResponse


def build_csv_response(
    *,
    rows: Sequence[dict[str, Any]],
    columns: list[tuple[str, str]],
    filename: str,
) -> HttpResponse:
    """
    Build a CSV HttpResponse from a list of row dicts.

    columns: list of (field_key, header_label) tuples.
    """
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([label for _, label in columns])

    for row in rows:
        writer.writerow([_format_value(row.get(key, "")) for key, label in columns])

    return response


def build_excel_response(
    *,
    rows: Sequence[dict[str, Any]],
    columns: list[tuple[str, str]],
    filename: str,
    sheet_name: str = "Export",
) -> HttpResponse:
    """
    Build an Excel (.xlsx) HttpResponse using openpyxl.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_format_value(row.get(key, "")))

    # Auto-width columns
    for col_idx, (_, label) in enumerate(columns, start=1):
        max_len = len(label)
        for row_idx in range(2, min(len(rows) + 2, 102)):  # sample first 100 rows
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _format_value(value: Any) -> str:
    """Format a value for export."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (list, dict)):
        import json
        return json.dumps(value)
    return str(value)
