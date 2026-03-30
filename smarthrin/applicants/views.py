"""Views for Applicant resource."""
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers
from rest_framework.decorators import action, api_view, authentication_classes, permission_classes
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from activities.models import Activity
from activities.services import log_activity, log_activity_for_request
from common.authentication import JWTRequestAuthentication
from common.mixins import TenantViewSetMixin
from common.pagination import StandardResultsPagination
from common.permissions import require_permission

from .filters import ApplicantFilterSet
from .models import Applicant
from .serializers import (
    ApplicantCreateSerializer,
    ApplicantDetailSerializer,
    ApplicantImportSerializer,
    ApplicantListSerializer,
)


# ------------------------------------------------------------------
# Standalone export view (explicit path avoids DefaultRouter issues)
# ------------------------------------------------------------------

@extend_schema(
    tags=["Applicants"],
    summary="Export applicants",
    description=(
        "Export filtered applicants as CSV or Excel. "
        "Supports the same query params as the list endpoint (source, email, skills, etc.). "
        "Use `export_format=xlsx` for Excel or `export_format=csv` (default) for CSV."
    ),
    parameters=[
        OpenApiParameter("export_format", OpenApiTypes.STR, enum=["csv", "xlsx"], description="Export format (default: csv)"),
        OpenApiParameter("source", OpenApiTypes.STR, description="Filter by source"),
        OpenApiParameter("search", OpenApiTypes.STR, description="Search in name/email"),
        OpenApiParameter("experience_years_gte", OpenApiTypes.NUMBER, description="Min experience years"),
        OpenApiParameter("experience_years_lte", OpenApiTypes.NUMBER, description="Max experience years"),
        OpenApiParameter("skills", OpenApiTypes.STR, description="Filter by skill"),
    ],
    responses={200: None},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.applicants.view")])
def export_applicants(request: Request):
    """Export filtered applicants to CSV or Excel."""
    import datetime
    import json
    from common.export import build_csv_response, build_excel_response

    qs = Applicant.objects.filter(tenant_id=request.tenant_id)

    # Honour permission_scope set by HasSmartHRPermission
    scope = getattr(request, "permission_scope", None)
    if scope == "own":
        qs = qs.filter(owner_user_id=request.user_id)

    filterset = ApplicantFilterSet(request.query_params, queryset=qs, request=request)
    if not filterset.is_valid():
        from rest_framework.exceptions import ValidationError
        raise ValidationError(filterset.errors)
    qs = filterset.qs

    export_format = request.query_params.get("export_format", "csv").lower()

    columns = [
        ("first_name", "First Name"),
        ("last_name", "Last Name"),
        ("email", "Email"),
        ("phone", "Phone"),
        ("source", "Source"),
        ("experience_years", "Experience (Years)"),
        ("current_company", "Current Company"),
        ("current_role", "Current Role"),
        ("skills", "Skills"),
        ("tags", "Tags"),
        ("linkedin_url", "LinkedIn URL"),
        ("portfolio_url", "Portfolio URL"),
        ("notes", "Notes"),
        ("created_at", "Created At"),
    ]

    rows = []
    for applicant in qs.iterator():
        rows.append({
            "first_name": applicant.first_name,
            "last_name": applicant.last_name,
            "email": applicant.email,
            "phone": applicant.phone,
            "source": applicant.source,
            "experience_years": str(applicant.experience_years) if applicant.experience_years is not None else "",
            "current_company": applicant.current_company,
            "current_role": applicant.current_role,
            "skills": ", ".join(applicant.skills or []),
            "tags": ", ".join(applicant.tags or []),
            "linkedin_url": applicant.linkedin_url,
            "portfolio_url": applicant.portfolio_url,
            "notes": applicant.notes,
            "created_at": applicant.created_at,
        })

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    if export_format == "xlsx":
        return build_excel_response(
            rows=rows,
            columns=columns,
            filename=f"applicants_{timestamp}.xlsx",
            sheet_name="Applicants",
        )
    return build_csv_response(
        rows=rows,
        columns=columns,
        filename=f"applicants_{timestamp}.csv",
    )


# ------------------------------------------------------------------
# Standalone import view
# ------------------------------------------------------------------

# All Applicant model fields that can be mapped from an Excel column.
# Virtual composite fields (prefixed with explanation) are also accepted;
# they are expanded into real DB fields during row processing.
IMPORTABLE_FIELDS: dict[str, str] = {
    "first_name": "First Name",
    "last_name": "Last Name",
    "full_name": "Full Name (splits into First + Last Name)",
    "email": "Email",
    "phone": "Phone",
    "resume_url": "Resume URL",
    "linkedin_url": "LinkedIn URL",
    "portfolio_url": "Portfolio URL",
    "skills": "Skills",
    "experience_years": "Experience (Years)",
    "current_company": "Current Company",
    "current_role": "Current Role",
    "notes": "Notes",
    "source": "Source",
    "tags": "Tags",
}

# Aliases for fuzzy auto-mapping of Excel column headers to DB fields.
# Keys are lowercase; values are the corresponding IMPORTABLE_FIELDS key.
_FIELD_ALIASES: dict[str, str] = {
    # full_name (virtual – split into first_name + last_name)
    "full name": "full_name",
    "full_name": "full_name",
    "fullname": "full_name",
    "name": "full_name",
    "candidate name": "full_name",
    "candidate": "full_name",
    "applicant name": "full_name",
    "applicant": "full_name",
    # first_name
    "first name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "given name": "first_name",
    "givenname": "first_name",
    "forename": "first_name",
    "prenom": "first_name",
    "prénom": "first_name",
    "nombre": "first_name",
    "first": "first_name",
    # last_name
    "last name": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "family name": "last_name",
    "familyname": "last_name",
    "nom": "last_name",
    "apellido": "last_name",
    "last": "last_name",
    # email
    "email": "email",
    "e-mail": "email",
    "email address": "email",
    "e-mail address": "email",
    "emailaddress": "email",
    "mail": "email",
    "contact email": "email",
    "work email": "email",
    "personal email": "email",
    # phone
    "phone": "phone",
    "phone number": "phone",
    "phonenumber": "phone",
    "telephone": "phone",
    "tel": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    "cell": "phone",
    "cell phone": "phone",
    "contact number": "phone",
    "contact": "phone",
    # resume_url
    "resume url": "resume_url",
    "resume link": "resume_url",
    "resume": "resume_url",
    "cv url": "resume_url",
    "cv link": "resume_url",
    "cv": "resume_url",
    # linkedin_url
    "linkedin url": "linkedin_url",
    "linkedin link": "linkedin_url",
    "linkedin": "linkedin_url",
    "linkedin profile": "linkedin_url",
    # portfolio_url
    "portfolio url": "portfolio_url",
    "portfolio link": "portfolio_url",
    "portfolio": "portfolio_url",
    "website": "portfolio_url",
    "personal website": "portfolio_url",
    # skills
    "skills": "skills",
    "skill": "skills",
    "skill set": "skills",
    "skillset": "skills",
    "competencies": "skills",
    "technologies": "skills",
    "tech stack": "skills",
    # experience_years
    "experience years": "experience_years",
    "experience (years)": "experience_years",
    "years of experience": "experience_years",
    "experience": "experience_years",
    "yoe": "experience_years",
    "exp": "experience_years",
    "total experience": "experience_years",
    # current_company
    "current company": "current_company",
    "company": "current_company",
    "employer": "current_company",
    "organization": "current_company",
    "organisation": "current_company",
    "current employer": "current_company",
    # current_role
    "current role": "current_role",
    "role": "current_role",
    "title": "current_role",
    "job title": "current_role",
    "position": "current_role",
    "designation": "current_role",
    "current title": "current_role",
    "current position": "current_role",
    # notes
    "notes": "notes",
    "note": "notes",
    "comments": "notes",
    "comment": "notes",
    "remarks": "notes",
    "remark": "notes",
    # source
    "source": "source",
    "lead source": "source",
    "channel": "source",
    "origin": "source",
    # tags
    "tags": "tags",
    "tag": "tags",
    "labels": "tags",
    "label": "tags",
    "categories": "tags",
    "category": "tags",
}


def _suggest_mapping(excel_columns: list[str]) -> dict[str, str]:
    """
    Return a best-effort mapping from Excel column names to DB field names.

    Uses exact alias matching (case-insensitive). Each DB field is mapped at
    most once (first match wins) to avoid ambiguous duplicates.

    Special rule: if both first_name and last_name are matched, full_name is
    excluded to avoid conflicts. Conversely, if only full_name matches, the
    separate first/last name columns won't be suggested.
    """
    suggested: dict[str, str] = {}
    used_fields: set[str] = set()

    # First pass: match all columns
    for col in excel_columns:
        normalised = col.strip().lower()
        if not normalised:
            continue
        db_field = _FIELD_ALIASES.get(normalised)
        if db_field and db_field not in used_fields:
            suggested[col] = db_field
            used_fields.add(db_field)

    # Resolve conflict: if first_name or last_name matched, drop full_name
    if ("first_name" in used_fields or "last_name" in used_fields) and "full_name" in used_fields:
        suggested = {k: v for k, v in suggested.items() if v != "full_name"}

    return suggested


@extend_schema(
    tags=["Applicants"],
    summary="Get importable fields",
    description=(
        "Returns the list of applicant database fields that can be mapped "
        "to Excel columns during import."
    ),
    responses={200: inline_serializer("ImportableFields", fields={
        "fields": serializers.DictField(child=serializers.CharField()),
    })},
)
@api_view(["GET"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.applicants.create")])
def import_fields(request: Request):
    """Return the list of importable fields for the frontend mapping UI."""
    return Response({"fields": IMPORTABLE_FIELDS})


@extend_schema(
    tags=["Applicants"],
    summary="Preview Excel columns",
    description=(
        "Upload an Excel file (.xlsx) and receive its column headers "
        "so the frontend can build a mapping UI."
    ),
    request={
        "multipart/form-data": inline_serializer("ImportPreviewRequest", fields={
            "file": serializers.FileField(),
        }),
    },
    responses={
        200: inline_serializer("ImportPreviewResponse", fields={
            "columns": serializers.ListField(child=serializers.CharField()),
            "sample_data": serializers.ListField(child=serializers.DictField()),
            "suggested_mapping": serializers.DictField(child=serializers.CharField()),
        }),
    },
)
@api_view(["POST"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.applicants.create")])
def import_preview(request: Request):
    """
    Accept an Excel file, return its column names and first 5 rows
    so the frontend can present a mapping UI.
    """
    import openpyxl

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"detail": "No file uploaded."}, status=400)

    if not uploaded_file.name.endswith((".xlsx", ".xls")):
        return Response({"detail": "Only .xlsx files are supported."}, status=400)

    try:
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        return Response({"detail": "Unable to read the uploaded file. Ensure it is a valid .xlsx file."}, status=400)

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return Response({"detail": "The uploaded file has no data."}, status=400)

    columns = [str(c).strip() if c is not None else f"Column {i+1}" for i, c in enumerate(header_row)]

    sample_data = []
    for _, row in zip(range(5), rows_iter):
        sample_data.append({
            columns[i]: (str(cell).strip() if cell is not None else "")
            for i, cell in enumerate(row)
            if i < len(columns)
        })

    wb.close()
    return Response({
        "columns": columns,
        "sample_data": sample_data,
        "suggested_mapping": _suggest_mapping(columns),
    })


@extend_schema(
    tags=["Applicants"],
    summary="Import applicants from Excel",
    description=(
        "Upload an Excel file and a field mapping to bulk-create applicants. "
        "No fields are required during import – only mapped columns are populated. "
        "Rows with duplicate emails (per tenant) are skipped. "
        "The response includes counts and per-row error details."
    ),
    request={
        "multipart/form-data": inline_serializer("ImportApplicantsRequest", fields={
            "file": serializers.FileField(help_text="The .xlsx file to import."),
            "mapping": serializers.CharField(
                help_text='JSON string mapping Excel column names to DB fields. '
                           'Example: {"Full Name": "first_name", "E-mail": "email"}',
            ),
        }),
    },
    responses={
        200: inline_serializer("ImportApplicantsResponse", fields={
            "total_rows": serializers.IntegerField(),
            "imported": serializers.IntegerField(),
            "skipped": serializers.IntegerField(),
            "errors": serializers.ListField(child=serializers.DictField()),
        }),
    },
)
@api_view(["POST"])
@authentication_classes([JWTRequestAuthentication])
@permission_classes([require_permission("smarthrin.applicants.create")])
def import_applicants(request: Request):
    """
    Bulk-import applicants from an uploaded Excel file.

    Expected multipart/form-data payload:
        file    – .xlsx file
        mapping – JSON string, e.g. {"Excel Col": "db_field", ...}
    """
    import json
    import uuid
    import openpyxl

    # ---- validate file -------------------------------------------------
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"detail": "No file uploaded."}, status=400)

    if not uploaded_file.name.endswith((".xlsx", ".xls")):
        return Response({"detail": "Only .xlsx files are supported."}, status=400)

    # ---- validate mapping ----------------------------------------------
    raw_mapping = request.data.get("mapping")
    if not raw_mapping:
        return Response({"detail": "Field mapping is required."}, status=400)

    try:
        mapping: dict = json.loads(raw_mapping) if isinstance(raw_mapping, str) else raw_mapping
    except (json.JSONDecodeError, TypeError):
        return Response({"detail": "Invalid mapping JSON."}, status=400)

    # Ensure every target field is a real importable field
    invalid_fields = [v for v in mapping.values() if v not in IMPORTABLE_FIELDS]
    if invalid_fields:
        return Response(
            {"detail": f"Invalid target fields: {', '.join(invalid_fields)}"},
            status=400,
        )

    # ---- read workbook --------------------------------------------------
    try:
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        return Response({"detail": "Unable to read the uploaded file."}, status=400)

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return Response({"detail": "The uploaded file has no data."}, status=400)

    columns = [str(c).strip() if c is not None else "" for c in header_row]

    # Build col-index → db_field lookup from the mapping
    col_to_field: dict[int, str] = {}
    for excel_col, db_field in mapping.items():
        try:
            idx = columns.index(excel_col)
            col_to_field[idx] = db_field
        except ValueError:
            wb.close()
            return Response(
                {"detail": f"Mapped Excel column '{excel_col}' not found in file headers."},
                status=400,
            )

    # ---- pre-fetch existing emails for this tenant (for skip logic) -----
    existing_emails: set[str] = set(
        Applicant.objects.filter(tenant_id=request.tenant_id)
        .values_list("email", flat=True)
    )

    # ---- iterate rows and build applicant dicts -------------------------
    imported = 0
    skipped = 0
    errors: list[dict] = []
    batch: list[dict] = []

    for row_num, row in enumerate(rows_iter, start=2):
        row_data: dict = {}
        for col_idx, db_field in col_to_field.items():
            cell_value = row[col_idx] if col_idx < len(row) else None
            if cell_value is None:
                continue

            cell_str = str(cell_value).strip()
            if not cell_str:
                continue

            # Convert comma-separated strings to lists for JSON fields
            if db_field in ("skills", "tags"):
                row_data[db_field] = [s.strip() for s in cell_str.split(",") if s.strip()]
            elif db_field == "experience_years":
                try:
                    row_data[db_field] = int(float(cell_str))
                except (ValueError, TypeError):
                    pass  # skip non-numeric, field stays empty
            elif db_field == "full_name":
                # Virtual field: split "Jane Doe" → first_name + last_name
                parts = cell_str.split(None, 1)
                row_data["first_name"] = parts[0]
                row_data["last_name"] = parts[1] if len(parts) > 1 else ""
            else:
                row_data[db_field] = cell_str

        # Skip completely empty rows
        if not row_data:
            continue

        # Force source to IMPORT unless the user explicitly mapped a source column
        if "source" not in col_to_field.values():
            row_data["source"] = "IMPORT"

        # Duplicate email check (skip row, don't error)
        email = row_data.get("email", "").lower()
        if email and email in existing_emails:
            skipped += 1
            continue

        # Validate through serializer (no required fields)
        ser = ApplicantImportSerializer(data=row_data, context={"request": request})
        if not ser.is_valid():
            errors.append({"row": row_num, "errors": ser.errors})
            continue

        batch.append(ser.validated_data)

        # Track the email so later rows in the same file won't duplicate it
        if email:
            existing_emails.add(email)

    wb.close()

    # ---- bulk create ----------------------------------------------------
    applicants_to_create = [
        Applicant(
            tenant_id=request.tenant_id,
            owner_user_id=request.user_id,
            **data,
        )
        for data in batch
    ]

    if applicants_to_create:
        Applicant.objects.bulk_create(applicants_to_create, ignore_conflicts=True)
        imported = len(applicants_to_create)

        # Log a single activity for the whole import
        log_activity(
            tenant_id=str(request.tenant_id),
            actor_user_id=str(request.user_id),
            actor_email=getattr(request, "email", ""),
            verb=Activity.Verb.CREATED,
            resource_type="Applicant",
            resource_id=str(uuid.uuid4()),
            resource_label=f"Bulk import of {imported} applicants",
            after={"action": "bulk_import", "count": imported},
        )

    total_rows = imported + skipped + len(errors)

    return Response({
        "total_rows": total_rows,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    })


@extend_schema_view(
    list=extend_schema(
        tags=["Applicants"],
        summary="List applicants",
        description="Returns paginated list of applicants for the current tenant.",
        parameters=[
            OpenApiParameter("source", OpenApiTypes.STR, enum=["MANUAL", "WEBSITE", "LINKEDIN", "REFERRAL", "IMPORT"],
                description="Filter by source"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Search in first_name, last_name, email"),
        ],
        responses={200: ApplicantListSerializer(many=True)},
    ),
    create=extend_schema(
        tags=["Applicants"],
        summary="Create an applicant",
        request=ApplicantCreateSerializer,
        responses={201: ApplicantDetailSerializer},
        examples=[
            OpenApiExample(
                "New applicant",
                value={
                    "first_name": "Alice", "last_name": "Johnson",
                    "email": "alice@example.com", "phone": "+14155550001",
                    "skills": ["Python", "Django"], "experience_years": 5,
                    "source": "LINKEDIN",
                },
            ),
        ],
    ),
    retrieve=extend_schema(
        tags=["Applicants"],
        summary="Get applicant details",
        responses={200: ApplicantDetailSerializer},
    ),
    update=extend_schema(
        tags=["Applicants"],
        summary="Update an applicant",
        request=ApplicantCreateSerializer,
        responses={200: ApplicantDetailSerializer},
    ),
    partial_update=extend_schema(
        tags=["Applicants"],
        summary="Partial update an applicant",
        request=ApplicantCreateSerializer,
        responses={200: ApplicantDetailSerializer},
    ),
    destroy=extend_schema(
        tags=["Applicants"],
        summary="Delete an applicant",
        responses={204: None},
    ),
)
class ApplicantViewSet(TenantViewSetMixin, ModelViewSet):
    """CRUD + extra actions for Applicant."""

    queryset = Applicant.objects.prefetch_related("applications__job")
    authentication_classes = [JWTRequestAuthentication]
    pagination_class = StandardResultsPagination
    filterset_class = ApplicantFilterSet
    search_fields = ["first_name", "last_name", "email"]
    ordering_fields = ["created_at", "updated_at", "first_name", "last_name"]

    def get_permissions(self):
        action_permission_map = {
            "list": require_permission("smarthrin.applicants.view"),
            "retrieve": require_permission("smarthrin.applicants.view"),
            "create": require_permission("smarthrin.applicants.create"),
            "update": require_permission("smarthrin.applicants.edit"),
            "partial_update": require_permission("smarthrin.applicants.edit"),
            "destroy": require_permission("smarthrin.applicants.delete"),
            "applications": require_permission("smarthrin.applications.view"),
        }
        perm_class = action_permission_map.get(
            self.action, require_permission("smarthrin.applicants.view")
        )
        return [perm_class()]

    def get_serializer_class(self):
        if self.action == "list":
            return ApplicantListSerializer
        if self.action in ("create", "update", "partial_update"):
            return ApplicantCreateSerializer
        return ApplicantDetailSerializer

    def perform_create(self, serializer):
        instance = serializer.save(
            tenant_id=self.request.tenant_id,
            owner_user_id=self.request.user_id,
        )
        log_activity_for_request(
            self.request,
            verb=Activity.Verb.CREATED,
            resource=instance,
            after={"email": instance.email},
        )

    def perform_update(self, serializer):
        instance = serializer.save(tenant_id=self.request.tenant_id)
        log_activity_for_request(
            self.request,
            verb=Activity.Verb.UPDATED,
            resource=instance,
        )

    @extend_schema(
        tags=["Applicants"],
        summary="List applicant's applications",
        description="Returns all job applications made by this applicant for the current tenant.",
        responses={200: inline_serializer("ApplicantApplicationsList", fields={
            "id": serializers.UUIDField(),
            "job_title": serializers.CharField(),
            "status": serializers.CharField(),
            "score": serializers.DecimalField(max_digits=5, decimal_places=2, allow_null=True),
            "created_at": serializers.DateTimeField(),
        }, many=True)},
    )
    @action(detail=True, methods=["get"], url_path="applications", url_name="applications")
    def applications(self, request, pk=None):
        """Return paginated applications for this applicant (tenant-filtered)."""
        from applications.models import Application
        from applications.serializers import ApplicationListSerializer

        applicant = self.get_object()
        qs = (
            Application.objects.filter(applicant=applicant, tenant_id=request.tenant_id)
            .select_related("applicant", "job")
            .order_by("-created_at")
        )

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = ApplicationListSerializer(
                page, many=True, context={"request": request}
            )
            return self.get_paginated_response(serializer.data)

        serializer = ApplicationListSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)
